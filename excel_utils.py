from __future__ import annotations
from typing import Dict, List, Tuple, Any, Optional
from io import BytesIO
from openpyxl import load_workbook, Workbook

def read_xlsx_rows(content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active
    header = []
    for c in ws[1]:
        header.append(str(c.value).strip() if c.value is not None else "")
    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip()=="" for v in r):
            continue
        d={}
        for i, col in enumerate(header):
            if not col:
                continue
            d[col]=r[i] if i < len(r) else None
        rows.append(d)
    return header, rows

def make_error_xlsx(original_header: List[str], bad_rows: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "errors"
    header = list(original_header)
    if "에러" not in header:
        header.append("에러")
    ws.append(header)
    for row in bad_rows:
        ws.append([row.get(h, "") for h in header])
    bio=BytesIO()
    wb.save(bio)
    return bio.getvalue()

def to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()

def to_int(v: Any) -> int:
    if v is None or str(v).strip()=="":
        raise ValueError("수량 없음")
    if isinstance(v, (int,)):
        return int(v)
    if isinstance(v, float):
        return int(v)
    s=str(v).strip()
    if s.isdigit() or (s.startswith("-") and s[1:].isdigit()):
        return int(s)
    # "10.0" 같은 값
    try:
        return int(float(s))
    except:
        raise ValueError("수량 형식 오류")
