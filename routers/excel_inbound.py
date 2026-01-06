from fastapi import APIRouter, UploadFile, File, Form, HTTPException
import openpyxl
import io

from app.db import upsert_inventory, add_history
from app.utils.excel_kor_columns import build_col_index, validate_required


router = APIRouter(prefix="/api/excel/inbound", tags=["excel-inbound"])


@router.post("")
async def excel_inbound(operator: str = Form(""), file: UploadFile = File(...)):
    """입고 엑셀 업로드 (한글 컬럼 고정)

    필수 컬럼: 창고, 로케이션, 브랜드, 품번, 품명, LOT, 규격, 수량
    선택 컬럼: 비고
    """
    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(status_code=400, detail="엑셀(.xlsx) 파일만 업로드 가능합니다.")

    data = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
    ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = build_col_index(headers)
    ok, missing = validate_required(idx)
    if not ok:
        raise HTTPException(status_code=400, detail=f"필수 컬럼 누락: {', '.join(missing)}")

    success = 0
    fail = 0
    errors = []

    for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        try:
            warehouse = str(row[idx["창고"]] or "").strip()
            location = str(row[idx["로케이션"]] or "").strip()
            brand = str(row[idx["브랜드"]] or "").strip()
            item_code = str(row[idx["품번"]] or "").strip()
            item_name = str(row[idx["품명"]] or "").strip()
            lot = str(row[idx["LOT"]] or "").strip()
            spec = str(row[idx["규격"]] or "").strip()
            qty_raw = row[idx["수량"]]
            note = ""
            if "비고" in idx and idx["비고"] < len(row):
                note = str(row[idx["비고"]] or "").strip()

            if not (warehouse and location and brand and item_code and item_name and lot and spec):
                raise ValueError("필수 값(창고/로케이션/브랜드/품번/품명/LOT/규격) 누락")

            qty = int(qty_raw)
            if qty <= 0:
                raise ValueError("수량은 1 이상")

            # 재고 증가
            upsert_inventory(warehouse, location, brand, item_code, item_name, lot, spec, qty, note)
            add_history("입고", warehouse, operator, brand, item_code, item_name, lot, spec, "", location, qty, note)
            success += 1
        except Exception as e:
            fail += 1
            errors.append({"row": r_i, "error": str(e)})

    return {"ok": True, "success": success, "fail": fail, "errors": errors[:50]}
