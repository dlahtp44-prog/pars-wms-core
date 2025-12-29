from fastapi import APIRouter, UploadFile, File
from openpyxl import load_workbook
from app.utils.excel_error import make_error_excel

router = APIRouter(prefix="/api/inbound", tags=["inbound"])

@router.post("/excel")
def inbound_excel(file: UploadFile = File(...)):
    wb = load_workbook(file.file)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    error_rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        location, code, name, lot, spec, brand, qty, memo = row

        if not location:
            error_rows.append((row, "로케이션 누락"))
            continue
        if not lot:
            error_rows.append((row, "LOT 누락"))
            continue
        if not qty or qty <= 0:
            error_rows.append((row, "수량은 1 이상"))
            continue

        # TODO: 정상 데이터 DB 저장 로직

    if error_rows:
        return make_error_excel(headers, error_rows, filename="inbound_error.xlsx")

    return {"result": "ok"}
