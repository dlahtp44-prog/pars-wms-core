from __future__ import annotations
from typing import Dict, List, Tuple, Any
from openpyxl import load_workbook

# Header synonyms mapping
SYNONYMS: Dict[str, str] = {
    "location": "location",
    "로케이션": "location",
    "위치": "location",

    "src_location": "src_location",
    "출발로케이션": "src_location",
    "출발": "src_location",
    "from": "src_location",

    "dst_location": "dst_location",
    "도착로케이션": "dst_location",
    "도착": "dst_location",
    "to": "dst_location",

    "item_code": "item_code",
    "품번": "item_code",
    "품목코드": "item_code",

    "item_name": "item_name",
    "품명": "item_name",
    "품목명": "item_name",

    "lot": "lot",
    "LOT": "lot",

    "spec": "spec",
    "규격": "spec",

    "brand": "brand",
    "브랜드": "brand",

    "qty": "qty",
    "수량": "qty",

    "note": "note",
    "비고": "note",
    "메모": "note",
}

def _norm(v: Any) -> str:
    return ("" if v is None else str(v)).strip()

def read_xlsx_rows(path: str) -> Tuple[List[str], List[List[Any]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(list(r))
    if not rows:
        return [], []
    header = [_norm(h) for h in rows[0]]
    data = rows[1:]
    # drop trailing empty rows
    cleaned = []
    for row in data:
        if all(_norm(c)=="" for c in row):
            continue
        cleaned.append(list(row))
    return header, cleaned

def map_columns(headers: List[str]) -> Dict[str, int]:
    colmap: Dict[str, int] = {}
    for i, h in enumerate(headers):
        key = SYNONYMS.get(h, SYNONYMS.get(h.lower(), None))
        if key:
            colmap[key] = i
    return colmap

def get(row: List[Any], colmap: Dict[str,int], key: str) -> str:
    i = colmap.get(key, None)
    if i is None or i >= len(row):
        return ""
    return _norm(row[i])

def get_int(row: List[Any], colmap: Dict[str,int], key: str) -> int:
    s = get(row, colmap, key)
    if s == "":
        return 0
    try:
        return int(float(s))
    except Exception:
        raise ValueError(f"{key} 값이 숫자가 아닙니다: {s}")



def parse_xlsx(path: str, mode: str = "inbound"):
    """Return list[dict] normalized rows. mode: inbound/outbound/move"""
    headers, data_rows = read_xlsx_rows(path)
    mapped = map_headers(headers)
    out = []
    for row in data_rows:
        item = {}
        for idx, key in mapped.items():
            item[key] = _norm(row[idx]) if idx < len(row) else ""
        # normalize qty
        if "qty" in item:
            try:
                item["qty"] = int(float(item["qty"])) if str(item["qty"]).strip() != "" else 0
            except Exception:
                item["qty"] = 0

        if mode == "move":
            req = ["src_location","dst_location","item_code","item_name","lot","spec","qty"]
        else:
            req = ["location","item_code","item_name","lot","spec","qty"]

        for k in req:
            if k not in item or str(item.get(k,"")).strip() == "" or (k=="qty" and int(item.get("qty",0))<=0):
                raise ValueError(f"엑셀 필수값 누락: {k}")
        out.append(item)
    return out
