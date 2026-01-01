import os, sqlite3
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from fastapi import UploadFile, HTTPException

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "WMS.db")

def _conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = _conn()
    cur = conn.cursor()
    # inventory: current stock by (location,item_code,lot,spec,brand)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS inventory (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        location TEXT NOT NULL,
        item_code TEXT NOT NULL,
        item_name TEXT NOT NULL,
        lot TEXT NOT NULL,
        spec TEXT NOT NULL,
        brand TEXT DEFAULT '',
        qty INTEGER NOT NULL DEFAULT 0,
        note TEXT DEFAULT '',
        updated_at TEXT NOT NULL
    );
    """)
    cur.execute("""
    CREATE UNIQUE INDEX IF NOT EXISTS ux_inventory_key
    ON inventory(location, item_code, lot, spec, brand);
    """)
    # history
    cur.execute("""
    CREATE TABLE IF NOT EXISTS history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ts TEXT NOT NULL,
        type TEXT NOT NULL, -- inbound/outbound/move
        location TEXT DEFAULT '',
        from_location TEXT DEFAULT '',
        to_location TEXT DEFAULT '',
        item_code TEXT NOT NULL,
        item_name TEXT NOT NULL,
        lot TEXT NOT NULL,
        spec TEXT NOT NULL,
        brand TEXT DEFAULT '',
        qty INTEGER NOT NULL,
        note TEXT DEFAULT ''
    );
    """)

# rollbacks: prevent double rollback (ref_history_id unique)
cur.execute("""
CREATE TABLE IF NOT EXISTS rollbacks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ref_history_id INTEGER NOT NULL UNIQUE,
    reason TEXT DEFAULT '',
    ts TEXT NOT NULL
);
""")
# manual log (파손/특이사항 등 수기입력)
cur.execute("""
CREATE TABLE IF NOT EXISTS manual_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ts TEXT NOT NULL,
    year INTEGER NOT NULL,
    month INTEGER NOT NULL,
    책임구분 TEXT NOT NULL,
    유형 TEXT NOT NULL,
    상황 TEXT NOT NULL,
    기타내용 TEXT DEFAULT '',
    등록일자 TEXT DEFAULT '',
    파손일자 TEXT DEFAULT '',
    담당자 TEXT DEFAULT '',
    수량 INTEGER DEFAULT 0,
    내용 TEXT DEFAULT ''
);
""")
    # calendar memo
    cur.execute("""
    CREATE TABLE IF NOT EXISTS calendar_memo (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        memo_date TEXT NOT NULL, -- YYYY-MM-DD
        author TEXT DEFAULT '',
        memo TEXT NOT NULL,
        created_at TEXT NOT NULL
    );
    """)
    conn.commit()
    conn.close()

def _now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def _upsert_inventory(location, item_code, item_name, lot, spec, brand, qty_delta, note=""):
    conn = _conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, qty FROM inventory
        WHERE location=? AND item_code=? AND lot=? AND spec=? AND brand=?
    """, (location, item_code, lot, spec, brand))
    row = cur.fetchone()
    updated_at = _now()
    if row:
        new_qty = int(row["qty"]) + int(qty_delta)
        if new_qty < 0:
            conn.close()
            raise HTTPException(status_code=400, detail=f"재고 부족: {location} / {item_code} / LOT {lot} ({row['qty']} -> {new_qty})")
        cur.execute("""
            UPDATE inventory
            SET item_name=?, qty=?, note=?, updated_at=?
            WHERE id=?
        """, (item_name, new_qty, note, updated_at, row["id"]))
    else:
        if int(qty_delta) < 0:
            conn.close()
            raise HTTPException(status_code=400, detail=f"재고 없음: {location} / {item_code} / LOT {lot}")
        cur.execute("""
            INSERT INTO inventory(location,item_code,item_name,lot,spec,brand,qty,note,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?)
        """, (location, item_code, item_name, lot, spec, brand, int(qty_delta), note, updated_at))
    conn.commit()
    conn.close()

def _add_history(type_, location, from_location, to_location, item_code, item_name, lot, spec, brand, qty, note):
    conn = _conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO history(ts,type,location,from_location,to_location,item_code,item_name,lot,spec,brand,qty,note)
        VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
    """, (_now(), type_, location, from_location, to_location, item_code, item_name, lot, spec, brand, int(qty), note))
    conn.commit()
    conn.close()

def add_inbound(location, item_code, item_name, lot, spec, brand, qty, note=""):
    _upsert_inventory(location, item_code, item_name, lot, spec, brand, int(qty), note)
    _add_history("inbound", location, "", "", item_code, item_name, lot, spec, brand, int(qty), note)

def add_outbound(location, item_code, item_name, lot, spec, brand, qty, note=""):
    _upsert_inventory(location, item_code, item_name, lot, spec, brand, -int(qty), note)
    _add_history("outbound", location, "", "", item_code, item_name, lot, spec, brand, int(qty), note)

def add_move(from_location, to_location, item_code, item_name, lot, spec, brand, qty, note=""):
    _upsert_inventory(from_location, item_code, item_name, lot, spec, brand, -int(qty), note)
    _upsert_inventory(to_location, item_code, item_name, lot, spec, brand, int(qty), note)
    _add_history("move", "", from_location, to_location, item_code, item_name, lot, spec, brand, int(qty), note)

def search_inventory(location: str = "", item_code: str = "", year: Optional[int]=None, month: Optional[int]=None) -> List[Dict[str, Any]]:
    conn = _conn()
    cur = conn.cursor()
    sql = "SELECT location,item_code,item_name,lot,spec,brand,qty,updated_at,note FROM inventory WHERE 1=1"
    params = []
    if location:
        sql += " AND location LIKE ?"
        params.append(f"%{location}%")
    if item_code:
        sql += " AND item_code LIKE ?"
        params.append(f"%{item_code}%")
    if year is not None:
        sql += " AND substr(updated_at,1,4)=?"
        params.append(f"{int(year):04d}")
    if month is not None:
        sql += " AND substr(updated_at,6,2)=?"
        params.append(f"{int(month):02d}")
    sql += " ORDER BY updated_at DESC"
    cur.execute(sql, params)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def get_history(limit: int = 200, year: Optional[int]=None, month: Optional[int]=None, htype: str="") -> List[Dict[str, Any]]:
    conn = _conn()
    cur = conn.cursor()
    sql = """
        SELECT id, ts,type,location,from_location,to_location,item_code,item_name,lot,spec,brand,qty,note
        FROM history WHERE 1=1
    """
    params=[]
    if htype:
        sql += " AND type=?"
        params.append(htype)
    if year is not None:
        sql += " AND substr(ts,1,4)=?"
        params.append(f"{int(year):04d}")
    if month is not None:
        sql += " AND substr(ts,6,2)=?"
        params.append(f"{int(month):02d}")
    sql += " ORDER BY id DESC LIMIT ?"
    params.append(int(limit))
    cur.execute(sql, params)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

# ---- Calendar memos ----
def upsert_calendar_memo(memo_date: str, author: str, memo: str):
    conn = _conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO calendar_memo(memo_date, author, memo, created_at)
        VALUES(?,?,?,?)
    """, (memo_date, author, memo, _now()))
    conn.commit()
    conn.close()

def get_calendar_memos_for_day(memo_date: str) -> List[Dict[str, Any]]:
    conn = _conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT author, memo, created_at FROM calendar_memo
        WHERE memo_date=?
        ORDER BY id DESC
    """, (memo_date,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def get_calendar_memos_for_month(year: int, month: int) -> Dict[str, List[Dict[str, Any]]]:
    # month range
    start = date(year, month, 1)
    if month == 12:
        end = date(year+1, 1, 1)
    else:
        end = date(year, month+1, 1)
    conn = _conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT memo_date, author, memo, created_at FROM calendar_memo
        WHERE memo_date >= ? AND memo_date < ?
        ORDER BY memo_date ASC, id DESC
    """, (start.isoformat(), end.isoformat()))
    out: Dict[str, List[Dict[str, Any]]] = {}
    for r in cur.fetchall():
        d = r["memo_date"]
        out.setdefault(d, []).append({"author": r["author"], "memo": r["memo"], "created_at": r["created_at"]})
    conn.close()
    return out

# ---- Excel helpers ----
def _norm(s: str) -> str:
    return (s or "").strip()

def _read_upload_bytes(file: UploadFile) -> bytes:
    return file.file.read()

def _sheet_rows(ws):
    # yields list of cell values for each row
    for row in ws.iter_rows(values_only=True):
        yield [c for c in row]

def _build_error_xlsx(rows: List[Dict[str, Any]], columns: List[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "errors"
    ws.append(columns)
    for r in rows:
        ws.append([r.get(c, "") for c in columns])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

async def parse_inbound_xlsx(file: UploadFile) -> Dict[str, Any]:
    """
    Required columns: 로케이션, 품번, 품명, LOT, 규격, 수량
    Optional: 브랜드, 비고
    """
    data = _read_upload_bytes(file)
    try:
        wb = load_workbook(filename=BytesIO(data))
    except Exception:
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx)만 업로드 가능합니다.")
    ws = wb.active
    header = [ _norm(str(x)) for x in next(_sheet_rows(ws)) ]
    col_map = {h: idx for idx, h in enumerate(header) if h}
    required = ["로케이션","품번","품명","LOT","규격","수량"]
    missing = [c for c in required if c not in col_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"필수 컬럼 누락: {', '.join(missing)}")
    ok, fail = 0, 0
    errors=[]
    for i, row in enumerate(list(_sheet_rows(ws))[1:], start=2):
        try:
            location=_norm(str(row[col_map["로케이션"]]))
            item_code=_norm(str(row[col_map["품번"]]))
            item_name=_norm(str(row[col_map["품명"]]))
            lot=_norm(str(row[col_map["LOT"]]))
            spec=_norm(str(row[col_map["규격"]]))
            qty=row[col_map["수량"]]
            brand=_norm(str(row[col_map["브랜드"]])) if "브랜드" in col_map else ""
            note=_norm(str(row[col_map["비고"]])) if "비고" in col_map else ""
            if not location or not item_code or not item_name or not lot or not spec:
                raise ValueError("필수값 공백")
            qty_i = int(qty)
            if qty_i <= 0:
                raise ValueError("수량은 1 이상")
            add_inbound(location,item_code,item_name,lot,spec,brand,qty_i,note)
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({
                "row": i, "error": str(e),
                "로케이션": row[col_map["로케이션"]] if col_map.get("로케이션") is not None else "",
                "품번": row[col_map["품번"]] if col_map.get("품번") is not None else "",
                "품명": row[col_map["품명"]] if col_map.get("품명") is not None else "",
                "LOT": row[col_map["LOT"]] if col_map.get("LOT") is not None else "",
                "규격": row[col_map["규격"]] if col_map.get("규격") is not None else "",
                "수량": row[col_map["수량"]] if col_map.get("수량") is not None else "",
                "브랜드": row[col_map["브랜드"]] if "브랜드" in col_map else "",
                "비고": row[col_map["비고"]] if "비고" in col_map else "",
            })
    error_xlsx=None
    if errors:
        cols=["row","error","로케이션","품번","품명","LOT","규격","수량","브랜드","비고"]
        error_xlsx=_build_error_xlsx(errors, cols)
    return {"ok": ok, "fail": fail, "errors": errors[:50], "error_xlsx_bytes": error_xlsx}

async def parse_outbound_xlsx(file: UploadFile) -> Dict[str, Any]:
    # same columns as inbound
    data=_read_upload_bytes(file)
    try:
        wb=load_workbook(filename=BytesIO(data))
    except Exception:
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx)만 업로드 가능합니다.")
    ws=wb.active
    header=[_norm(str(x)) for x in next(_sheet_rows(ws))]
    col_map={h: idx for idx,h in enumerate(header) if h}
    required=["로케이션","품번","품명","LOT","규격","수량"]
    missing=[c for c in required if c not in col_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"필수 컬럼 누락: {', '.join(missing)}")
    ok=fail=0
    errors=[]
    for i,row in enumerate(list(_sheet_rows(ws))[1:], start=2):
        try:
            location=_norm(str(row[col_map["로케이션"]]))
            item_code=_norm(str(row[col_map["품번"]]))
            item_name=_norm(str(row[col_map["품명"]]))
            lot=_norm(str(row[col_map["LOT"]]))
            spec=_norm(str(row[col_map["규격"]]))
            qty=row[col_map["수량"]]
            brand=_norm(str(row[col_map["브랜드"]])) if "브랜드" in col_map else ""
            note=_norm(str(row[col_map["비고"]])) if "비고" in col_map else ""
            qty_i=int(qty)
            if qty_i<=0:
                raise ValueError("수량은 1 이상")
            add_outbound(location,item_code,item_name,lot,spec,brand,qty_i,note)
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({"row": i, "error": str(e)})
    error_xlsx=None
    if errors:
        cols=["row","error"]
        error_xlsx=_build_error_xlsx(errors, cols)
    return {"ok": ok, "fail": fail, "errors": errors[:50], "error_xlsx_bytes": error_xlsx}

async def parse_move_xlsx(file: UploadFile) -> Dict[str, Any]:
    """
    Required: 출발로케이션, 도착로케이션, 품번, 품명, LOT, 규격, 수량
    Optional: 브랜드, 비고
    """
    data=_read_upload_bytes(file)
    try:
        wb=load_workbook(filename=BytesIO(data))
    except Exception:
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx)만 업로드 가능합니다.")
    ws=wb.active
    header=[_norm(str(x)) for x in next(_sheet_rows(ws))]
    col_map={h: idx for idx,h in enumerate(header) if h}
    required=["출발로케이션","도착로케이션","품번","품명","LOT","규격","수량"]
    missing=[c for c in required if c not in col_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"필수 컬럼 누락: {', '.join(missing)}")
    ok=fail=0
    errors=[]
    for i,row in enumerate(list(_sheet_rows(ws))[1:], start=2):
        try:
            from_loc=_norm(str(row[col_map["출발로케이션"]]))
            to_loc=_norm(str(row[col_map["도착로케이션"]]))
            item_code=_norm(str(row[col_map["품번"]]))
            item_name=_norm(str(row[col_map["품명"]]))
            lot=_norm(str(row[col_map["LOT"]]))
            spec=_norm(str(row[col_map["규격"]]))
            qty=row[col_map["수량"]]
            brand=_norm(str(row[col_map["브랜드"]])) if "브랜드" in col_map else ""
            note=_norm(str(row[col_map["비고"]])) if "비고" in col_map else ""
            qty_i=int(qty)
            if qty_i<=0:
                raise ValueError("수량은 1 이상")
            add_move(from_loc,to_loc,item_code,item_name,lot,spec,brand,qty_i,note)
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({"row": i, "error": str(e)})
    error_xlsx=None
    if errors:
        cols=["row","error"]
        error_xlsx=_build_error_xlsx(errors, cols)
    return {"ok": ok, "fail": fail, "errors": errors[:50], "error_xlsx_bytes": error_xlsx}

def inventory_to_xlsx(rows: List[Dict[str, Any]]) -> bytes:
    wb=Workbook()
    ws=wb.active
    ws.title="inventory"
    cols=["location","item_code","item_name","lot","spec","brand","qty","updated_at","note"]
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c,"") for c in cols])
    bio=BytesIO()
    wb.save(bio)
    return bio.getvalue()

def history_to_xlsx(rows: List[Dict[str, Any]]) -> bytes:
    wb=Workbook()
    ws=wb.active
    ws.title="history"
    cols=["ts","type","location","from_location","to_location","item_code","item_name","lot","spec","brand","qty","note"]
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c,"") for c in cols])
    bio=BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =========================
# MANUAL LOG (수기 입력)
# =========================
def add_manual_log(책임구분:str, 유형:str, 상황:str, 기타내용:str, 등록일자:str, 파손일자:str, 담당자:str, 수량:int, 내용:str):
    ts = _now()
    y, m = _year_month(ts)
    conn=_conn()
    cur=conn.cursor()
    cur.execute("""
        INSERT INTO manual_log(ts,year,month,책임구분,유형,상황,기타내용,등록일자,파손일자,담당자,수량,내용)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (ts,y,m,책임구분,유형,상황,기타내용 or "",등록일자 or "",파손일자 or "",담당자 or "",int(수량 or 0),내용 or ""))
    conn.commit()
    conn.close()

def get_manual_logs(year:int|None=None, month:int|None=None):
    conn=_conn()
    cur=conn.cursor()
    q="SELECT * FROM manual_log"
    params=[]
    cond=[]
    if year is not None:
        cond.append("year=?"); params.append(int(year))
    if month is not None:
        cond.append("month=?"); params.append(int(month))
    if cond:
        q += " WHERE " + " AND ".join(cond)
    q += " ORDER BY ts DESC"
    cur.execute(q, params)
    rows=[dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def manual_to_xlsx(year:int|None=None, month:int|None=None) -> bytes:
    rows = get_manual_logs(year, month)
    wb=Workbook()
    ws=wb.active
    ws.title="manual_log"
    headers=["ts","책임구분","유형","상황","기타내용","등록일자","파손일자","담당자","수량","내용"]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h,"") for h in headers])
    _autosize(ws)
    bio=BytesIO()
    wb.save(bio)
    return bio.getvalue()

# =========================
# ROLLBACK
# =========================
def _is_rolled_back(ref_history_id:int)->bool:
    conn=_conn()
    cur=conn.cursor()
    cur.execute("SELECT 1 FROM rollbacks WHERE ref_history_id=?", (int(ref_history_id),))
    ok = cur.fetchone() is not None
    conn.close()
    return ok

def rollback_history(ref_history_id:int, reason:str=""):
    if _is_rolled_back(ref_history_id):
        raise HTTPException(status_code=400, detail="이미 롤백 처리된 건입니다.")
    conn=_conn()
    cur=conn.cursor()
    cur.execute("SELECT * FROM history WHERE id=?", (int(ref_history_id),))
    h=cur.fetchone()
    if not h:
        conn.close()
        raise HTTPException(status_code=404, detail="원본 이력을 찾을 수 없습니다.")
    h=dict(h)
    if h["type"]=="rollback":
        conn.close()
        raise HTTPException(status_code=400, detail="롤백 건은 다시 롤백할 수 없습니다.")
    # apply inverse
    if h["type"]=="inbound":
        _upsert_inventory(h["location"], h["item_code"], h["item_name"], h["lot"], h["spec"], h.get("brand",""), -int(h["qty"]), reason)
        _add_history("rollback", h["location"], "", "", h["item_code"], h["item_name"], h["lot"], h["spec"], h.get("brand",""), int(h["qty"]), f"rollback_of={ref_history_id} {reason}".strip())
    elif h["type"]=="outbound":
        _upsert_inventory(h["location"], h["item_code"], h["item_name"], h["lot"], h["spec"], h.get("brand",""), int(h["qty"]), reason)
        _add_history("rollback", h["location"], "", "", h["item_code"], h["item_name"], h["lot"], h["spec"], h.get("brand",""), int(h["qty"]), f"rollback_of={ref_history_id} {reason}".strip())
    elif h["type"]=="move":
        # move inverse: to -> from
        _upsert_inventory(h["to_location"], h["item_code"], h["item_name"], h["lot"], h["spec"], h.get("brand",""), -int(h["qty"]), reason)
        _upsert_inventory(h["from_location"], h["item_code"], h["item_name"], h["lot"], h["spec"], h.get("brand",""), int(h["qty"]), reason)
        _add_history("rollback", "", h["to_location"], h["from_location"], h["item_code"], h["item_name"], h["lot"], h["spec"], h.get("brand",""), int(h["qty"]), f"rollback_of={ref_history_id} {reason}".strip())
    else:
        conn.close()
        raise HTTPException(status_code=400, detail=f"롤백 지원하지 않는 유형: {h['type']}")
    # mark rollback
    cur.execute("INSERT INTO rollbacks(ref_history_id, reason, ts) VALUES (?,?,?)", (int(ref_history_id), reason or "", _now()))
    conn.commit()
    conn.close()
    return {"result":"ok", "ref_history_id": int(ref_history_id)}
