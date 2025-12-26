from fastapi import APIRouter, Request, UploadFile, File, Depends
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
import os, time

from app.db import get_db
from app.core.auth import admin_required
from app.core.downloads import register_file
from app.core.excel_utils import normalize_headers, write_fail_xlsx, summarize

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

HEADERS = ["창고","출발로케이션","도착로케이션","품번","품명","LOT","규격","수량","비고"]

@router.get("/엑셀-이동", response_class=HTMLResponse)
def page(request: Request, _admin: str = Depends(admin_required)):
    return templates.TemplateResponse("excel_move.html", {"request": request})

@router.post("/엑셀-이동", response_class=HTMLResponse)
def upload(request: Request, file: UploadFile = File(...), _admin: str = Depends(admin_required)):
    t0 = time.time()
    wb = load_workbook(file.file, read_only=True, data_only=True)
    ws = wb.active

    headers = normalize_headers([c.value for c in ws[1]])
    errors = []
    failed_rows = []

    if headers != HEADERS:
        errors.append("엑셀 헤더가 양식과 다릅니다. (창고, 출발로케이션, 도착로케이션, 품번, 품명, LOT, 규격, 수량, 비고)")
        stats = summarize(0, 0, 0, time.time() - t0)
        return templates.TemplateResponse("excel_result.html", {
            "request": request,
            "success": [],
            "errors": errors,
            "retry_url": "/엑셀-이동",
            "stats": stats,
            "download_url": None
        })

    conn = get_db()
    cur = conn.cursor()

    total = ok = fail = 0

    sql_sel = "SELECT 수량 FROM 재고 WHERE 창고=? AND 로케이션=? AND 품번=? AND LOT=?"
    sql_deduct = """UPDATE 재고 SET 수량=수량-?, updated_at=datetime('now','localtime')
                   WHERE 창고=? AND 로케이션=? AND 품번=? AND LOT=?"""
    sql_del = "DELETE FROM 재고 WHERE 창고=? AND 로케이션=? AND 품번=? AND LOT=? AND 수량<=0"

    sql_upsert = """INSERT INTO 재고(창고, 로케이션, 품번, 품명, LOT, 규격, 수량, 비고)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ON CONFLICT(창고, 로케이션, 품번, LOT)
    DO UPDATE SET
      수량 = 재고.수량 + excluded.수량,
      품명 = excluded.품명,
      규격 = excluded.규격,
      비고 = excluded.비고,
      updated_at = datetime('now','localtime')"""

    sql_hist = """INSERT INTO 이력(구분, 창고, 품번, LOT, 출발로케이션, 도착로케이션, 수량, 비고)
    VALUES ('이동', ?, ?, ?, ?, ?, ?, ?)"""

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        total += 1
        try:
            창고, 출발, 도착, 품번, 품명, LOT, 규격, 수량, 비고 = row
            if not all([창고, 출발, 도착, 품번, 품명, LOT, 규격]):
                raise ValueError("필수값 누락")
            if int(수량) <= 0:
                raise ValueError("수량은 1 이상")
            if str(출발).strip() == str(도착).strip():
                raise ValueError("출발/도착 로케이션이 같습니다.")
            비고 = "" if 비고 is None else str(비고).strip()

            cur.execute("SAVEPOINT sp")
            cur.execute(sql_sel, (str(창고).strip(), str(출발).strip(), str(품번).strip(), str(LOT).strip()))
            r = cur.fetchone()
            if not r:
                raise ValueError("출발 로케이션에 재고가 없습니다.")
            if int(r["수량"]) < int(수량):
                raise ValueError("출발지 재고 수량이 부족합니다.")

            cur.execute(sql_deduct, (int(수량), str(창고).strip(), str(출발).strip(), str(품번).strip(), str(LOT).strip()))
            cur.execute(sql_del, (str(창고).strip(), str(출발).strip(), str(품번).strip(), str(LOT).strip()))
            cur.execute(sql_upsert, (str(창고).strip(), str(도착).strip(), str(품번).strip(), str(품명).strip(),
                                     str(LOT).strip(), str(규격).strip(), int(수량), 비고))
            cur.execute(sql_hist, (str(창고).strip(), str(품번).strip(), str(LOT).strip(), str(출발).strip(), str(도착).strip(), int(수량), 비고))
            cur.execute("RELEASE sp")

            ok += 1
        except Exception as e:
            fail += 1
            try:
                cur.execute("ROLLBACK TO sp")
                cur.execute("RELEASE sp")
            except Exception:
                pass
            reason = str(e)
            failed_rows.append((idx, list(row), reason))
            errors.append(f"{idx}행 실패: {reason}")

    conn.commit()
    conn.close()

    download_url = None
    if failed_rows:
        out_path = os.path.join("/tmp", f"failed_move_{int(time.time())}.xlsx")
        write_fail_xlsx(HEADERS, failed_rows, out_path)
        token = register_file(out_path)
        download_url = f"/download/{token}"

    stats = summarize(total, ok, fail, time.time() - t0)
    return templates.TemplateResponse("excel_result.html", {
        "request": request,
        "success": [f"성공 {ok}건"],
        "errors": errors,
        "retry_url": "/엑셀-이동",
        "stats": stats,
        "download_url": download_url
    })
