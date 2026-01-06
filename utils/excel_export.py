from __future__ import annotations

from io import BytesIO
from typing import Iterable, Mapping, Sequence, Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def rows_to_xlsx_bytes(
    rows: Iterable[Mapping[str, Any]],
    columns: Sequence[tuple[str, str]],
    sheet_name: str = "Sheet1",
) -> bytes:
    """rows(list[dict]) -> xlsx bytes.
    columns: [(key, header), ...]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # header
    for c, (_, header) in enumerate(columns, start=1):
        ws.cell(row=1, column=c, value=header)

    # body
    for r_idx, r in enumerate(rows, start=2):
        for c, (key, _) in enumerate(columns, start=1):
            ws.cell(row=r_idx, column=c, value=r.get(key, ""))

    # simple auto-width
    for c in range(1, len(columns) + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = max(10, min(35, ws.column_dimensions[col_letter].width or 10))

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
