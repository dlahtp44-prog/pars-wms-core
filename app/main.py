
import os
import sqlite3
import io
import datetime
import re
from typing import Optional

from fastapi import FastAPI, Request, Form, Query, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, Response, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware

from openpyxl import Workbook, load_workbook

APP_TITLE = "PARS WMS CORE"
ADMIN_PIN = os.getenv("ADMIN_PIN", "0000")
DB_PATH = os.getenv("DB_PATH", "wms.db")

app = FastAPI(title=APP_TITLE, version="2.0.0")
templates = Jinja2Templates(directory="app/templates")
app.mount("/static", StaticFiles(directory="app/static"), name="static")
app.add_middleware(SessionMiddleware, secret_key=os.getenv("SESSION_SECRET", "pars-wms-session-secret"))

def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    db = get_db()
    c = db.cursor()
    c.execute("""
    CREATE TABLE IF NOT EXISTS stock (
        item TEXT NOT NULL,
        item_name TEXT NOT NULL DEFAULT '',
        lot TEXT NOT NULL DEFAULT '',
        spec TEXT NOT NULL DEFAULT '',
        location TEXT NOT NULL,
        qty INTEGER NOT NULL,
        updated_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
        PRIMARY KEY(item, lot, location)
    )""")
    c.execute("""
    CREATE TABLE IF NOT EXISTS history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        type TEXT NOT NULL,
        item TEXT NOT NULL,
        item_name TEXT NOT NULL DEFAULT '',
        lot TEXT NOT NULL DEFAULT '',
        spec TEXT NOT NULL DEFAULT '',
        qty INTEGER NOT NULL,
        src TEXT NOT NULL DEFAULT '',
        dst TEXT NOT NULL DEFAULT '',
        memo TEXT NOT NULL DEFAULT '',
        ts TEXT NOT NULL DEFAULT (datetime('now','localtime'))
    )""")
    c.execute("""
    CREATE TABLE IF NOT EXISTS calendar_memo (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT NOT NULL,
        memo TEXT NOT NULL,
        author TEXT NOT NULL DEFAULT '',
        updated_at TEXT NOT NULL DEFAULT (datetime('now','localtime'))
    )""")

    # --- Schema migration (add brand/memo columns if missing) ---
    def _ensure_column(table: str, col: str, coldef: str):
        cols = [r[1] for r in c.execute(f"PRAGMA table_info({table})").fetchall()]
        if col not in cols:
            c.execute(f"ALTER TABLE {table} ADD COLUMN {col} {coldef}")

    _ensure_column("stock", "brand", "TEXT NOT NULL DEFAULT ''")
    _ensure_column("stock", "memo", "TEXT NOT NULL DEFAULT ''")
    _ensure_column("history", "brand", "TEXT NOT NULL DEFAULT ''")

    db.commit()
    db.close()

@app.on_event("startup")
def _startup():
    init_db()

# --- Quiet icon routes to avoid noise 404s ---
@app.get("/favicon.ico")
def favicon():
    return Response(status_code=204)

@app.get("/apple-touch-icon.png")
@app.get("/apple-touch-icon-precomposed.png")
def apple_icons():
    return Response(status_code=204)

# -----------------------
# Helpers: session QR
# -----------------------
def set_qr(request: Request, key: str, value: str):
    request.session.setdefault("qr", {})
    request.session["qr"][key] = value

def pop_qr(request: Request, key: str) -> str:
    qr = request.session.get("qr", {})
    v = qr.get(key, "")
    if key in qr:
        del qr[key]
        request.session["qr"] = qr
    return v

# -----------------------
# UI: Mobile Menu
# -----------------------
@app.get("/", response_class=HTMLResponse)
def mobile_menu(request: Request):
    return templates.TemplateResponse("menu.html", {"request": request, "title": APP_TITLE})

# Korean alias routes (fixes your logs: /입고 etc.)
@app.get("/입고")
def k_in():
    return RedirectResponse("/page/inbound", status_code=307)

@app.get("/출고")
def k_out():
    return RedirectResponse("/page/outbound", status_code=307)

@app.get("/이동")
def k_move():
    return RedirectResponse("/page/move", status_code=307)

@app.get("/재고")
def k_inv():
    return RedirectResponse("/page/inventory", status_code=307)

@app.get("/이력")
def k_hist():
    return RedirectResponse("/page/history", status_code=307)

@app.get("/달력")
def k_cal():
    # 달력 메뉴는 월간 달력으로 진입
    return RedirectResponse("/page/calendar/month", status_code=307)

@app.get("/관리자")
def k_admin():
    return RedirectResponse("/page/admin", status_code=307)

# -----------------------
# UI Pages
# -----------------------
@app.get("/page/inbound", response_class=HTMLResponse)
def page_inbound(request: Request):
    qr_loc = pop_qr(request, "in_loc")
    qr_item = pop_qr(request, "in_item")
    return templates.TemplateResponse("inbound.html", {"request": request, "qr_loc": qr_loc, "qr_item": qr_item})

@app.get("/page/outbound", response_class=HTMLResponse)
def page_outbound(request: Request):
    qr_loc = pop_qr(request, "out_loc")
    qr_item = pop_qr(request, "out_item")
    return templates.TemplateResponse("outbound.html", {"request": request, "qr_loc": qr_loc, "qr_item": qr_item})

@app.get("/page/move", response_class=HTMLResponse)
def page_move(request: Request):
    src = pop_qr(request, "mv_src")
    dst = pop_qr(request, "mv_dst")
    item = pop_qr(request, "mv_item")
    return templates.TemplateResponse("move.html", {"request": request, "src": src, "dst": dst, "item": item})

@app.get("/page/inventory", response_class=HTMLResponse)
def page_inventory(request: Request, location: str = "", item: str = ""):
    db = get_db()
    q = "SELECT item,item_name,lot,spec,brand,location,qty,updated_at,memo FROM stock WHERE 1=1"
    args = []
    if location:
        q += " AND location LIKE ?"
        args.append(f"%{location}%")
    if item:
        q += " AND item LIKE ?"
        args.append(f"%{item}%")
    q += " ORDER BY location, item"
    rows = db.execute(q, args).fetchall()
    db.close()
    return templates.TemplateResponse("inventory.html", {"request": request, "rows": rows, "location": location, "item": item})

@app.get("/page/history", response_class=HTMLResponse)
def page_history(request: Request, limit: int = 200):
    db = get_db()
    rows = db.execute("SELECT type,item,item_name,lot,spec,brand,qty,src,dst,memo,ts FROM history ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    db.close()
    return templates.TemplateResponse("history.html", {"request": request, "rows": rows, "limit": limit})

@app.get("/page/calendar", response_class=HTMLResponse)
def page_calendar(request: Request, date: str = ""):
    db = get_db()
    if date:
        rows = db.execute("SELECT id,date,memo,author,updated_at FROM calendar_memo WHERE date=? ORDER BY id DESC", (date,)).fetchall()
    else:
        rows = db.execute("SELECT id,date,memo,author,updated_at FROM calendar_memo ORDER BY date DESC, id DESC LIMIT 200").fetchall()
    db.close()
    return templates.TemplateResponse("calendar.html", {"request": request, "rows": rows, "date": date})

@app.get("/page/admin", response_class=HTMLResponse)
def page_admin(request: Request, pin: str = Query("", alias="pin")):
    authed = request.session.get("admin", False)
    if pin and pin == ADMIN_PIN:
        request.session["admin"] = True
        authed = True
    return templates.TemplateResponse("admin.html", {"request": request, "authed": authed})

# -----------------------
# QR Camera Page
# -----------------------
@app.get("/m/qr", response_class=HTMLResponse)
def page_qr(request: Request, mode: str = "in_loc"):
    # mode: in_loc, in_item, out_loc, out_item, mv_src, mv_dst, mv_item
    return templates.TemplateResponse("qr.html", {"request": request, "mode": mode})

@app.post("/m/qr/save")
def qr_save(request: Request, mode: str = Form(...), value: str = Form(...)):
    value = (value or "").strip()
    if not value:
        return RedirectResponse(f"/m/qr?mode={mode}", status_code=303)
    set_qr(request, mode, value)
    # back mapping
    if mode.startswith("in_"):
        return RedirectResponse("/page/inbound", status_code=303)
    if mode.startswith("out_"):
        return RedirectResponse("/page/outbound", status_code=303)
    if mode.startswith("mv_"):
        return RedirectResponse("/page/move", status_code=303)
    return RedirectResponse("/", status_code=303)

# -----------------------
# APIs (DB real)
# -----------------------
@app.post("/api/inbound")
def api_inbound(
    item: str = Form(...),
    item_name: str = Form(""),
    lot: str = Form(""),
    spec: str = Form(""),
    brand: str = Form(""),
    location: str = Form(...),
    qty: int = Form(...),
    memo: str = Form(""),
):
    item=item.strip(); location=location.strip()
    db=get_db()
    # upsert
    db.execute("""
    INSERT INTO stock(item,item_name,lot,spec,brand,location,qty,memo)
    VALUES(?,?,?,?,?,?,?,?)
    ON CONFLICT(item,lot,location) DO UPDATE SET
      qty = qty + excluded.qty,
      item_name = excluded.item_name,
      spec = excluded.spec,
      brand = excluded.brand,
      memo = excluded.memo,
      updated_at = datetime('now','localtime')
    """, (item,item_name,lot,spec,brand,location,int(qty),memo))
    db.execute("INSERT INTO history(type,item,item_name,lot,spec,brand,qty,src,dst,memo) VALUES('IN',?,?,?,?,?,?,?,?,?)",
               (item,item_name,lot,spec,brand,int(qty),"",location,memo))
    db.commit(); db.close()
    return RedirectResponse("/page/inbound", status_code=303)

@app.post("/api/outbound")
def api_outbound(
    item: str = Form(...),
    item_name: str = Form(""),
    lot: str = Form(""),
    spec: str = Form(""),
    brand: str = Form(""),
    location: str = Form(...),
    qty: int = Form(...),
    memo: str = Form(""),
):
    item=item.strip(); location=location.strip()
    db=get_db()
    # subtract if exists
    row = db.execute("SELECT qty FROM stock WHERE item=? AND lot=? AND location=?", (item,lot,location)).fetchone()
    cur_qty = int(row["qty"]) if row else 0
    if cur_qty < int(qty):
        db.close()
        return templates.TemplateResponse("message.html", {"request": Request, "title":"출고 오류", "msg":"재고가 부족합니다."}, status_code=400)
    db.execute("UPDATE stock SET qty = qty - ?, updated_at=datetime('now','localtime') WHERE item=? AND lot=? AND location=?",
               (int(qty),item,lot,location))
    db.execute("INSERT INTO history(type,item,item_name,lot,spec,brand,qty,src,dst,memo) VALUES('OUT',?,?,?,?,?,?,?,?,?)",
               (item,item_name,lot,spec,brand,int(qty),location,"",memo))
    db.commit(); db.close()
    return RedirectResponse("/page/outbound", status_code=303)

@app.post("/api/move")
def api_move(
    item: str = Form(...),
    item_name: str = Form(""),
    lot: str = Form(""),
    spec: str = Form(""),
    brand: str = Form(""),
    src: str = Form(...),
    dst: str = Form(...),
    qty: int = Form(...),
    memo: str = Form(""),
):
    item=item.strip(); src=src.strip(); dst=dst.strip()
    db=get_db()
    row = db.execute("SELECT qty FROM stock WHERE item=? AND lot=? AND location=?", (item,lot,src)).fetchone()
    cur_qty = int(row["qty"]) if row else 0
    if cur_qty < int(qty):
        db.close()
        return JSONResponse({"error":"재고 부족"}, status_code=400)
    # subtract src
    db.execute("UPDATE stock SET qty = qty - ?, updated_at=datetime('now','localtime') WHERE item=? AND lot=? AND location=?",
               (int(qty),item,lot,src))
    # add dst upsert
    db.execute("""
    INSERT INTO stock(item,item_name,lot,spec,brand,location,qty,memo)
    VALUES(?,?,?,?,?,?,?,?)
    ON CONFLICT(item,lot,location) DO UPDATE SET
      qty = qty + excluded.qty,
      item_name = excluded.item_name,
      spec = excluded.spec,
      brand = excluded.brand,
      memo = excluded.memo,
      updated_at = datetime('now','localtime')
    """, (item,item_name,lot,spec,dst,int(qty)))
    db.execute("INSERT INTO history(type,item,item_name,lot,spec,brand,qty,src,dst,memo) VALUES('MOVE',?,?,?,?,?,?,?,?,?)",
               (item,item_name,lot,spec,brand,int(qty),src,dst,memo))
    db.commit(); db.close()
    return RedirectResponse("/page/move", status_code=303)


# -----------------------
# Excel helpers / APIs
# -----------------------
def _xlsx_rows(file_bytes: bytes):
    """Read first sheet of xlsx and yield dicts by header row.
    Supports Korean/English headers.
    """
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    def norm(h: str) -> str:
        return re.sub(r"\s+", "", h.lower())
    # header key map
    mapping = {
        "로케이션":"location",
        "위치":"location",
        "location":"location",
        "품번":"item",
        "품목":"item",
        "item":"item",
        "품명":"item_name",
        "itemname":"item_name",
        "제품명":"item_name",
        "lot":"lot",
        "규격":"spec",
        "spec":"spec",
        "브랜드":"brand",
        "brand":"brand",
        "수량":"qty",
        "qty":"qty",
        "비고":"memo",
        "메모":"memo",
        "memo":"memo",
        "출발":"src",
        "from":"src",
        "src":"src",
        "도착":"dst",
        "to":"dst",
        "dst":"dst",
    }
    keys=[]
    for h in header:
        k = mapping.get(h, mapping.get(h.strip(), None))
        if not k:
            k = mapping.get(h.replace(" ", ""), None)
        if not k:
            k = mapping.get(h, None)
        if not k:
            # try normalized
            k = mapping.get(norm(h), None)
        keys.append(k)

    out=[]
    for r in rows[1:]:
        if r is None: 
            continue
        d={}
        empty=True
        for k,v in zip(keys, r):
            if k is None:
                continue
            if v is None:
                d[k]=""
            else:
                d[k]=str(v).strip() if isinstance(v,str) else v
            if d[k] not in ("", None):
                empty=False
        if not d or empty:
            continue
        out.append(d)
    return out

def _xlsx_response(filename: str, headers: list, data_rows: list):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in data_rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    return Response(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.post("/api/inbound/excel")
async def api_inbound_excel(file: UploadFile = File(...)):
    data = await file.read()
    rows = _xlsx_rows(data)
    db = get_db()
    for d in rows:
        location = str(d.get("location","")).strip()
        item = str(d.get("item","")).strip()
        if not location or not item:
            continue
        item_name = str(d.get("item_name","") or "")
        lot = str(d.get("lot","") or "")
        spec = str(d.get("spec","") or "")
        brand = str(d.get("brand","") or "")
        memo = str(d.get("memo","") or "")
        qty = int(float(d.get("qty",0) or 0))
        if qty <= 0:
            continue
        db.execute("""
        INSERT INTO stock(item,item_name,lot,spec,brand,location,qty,memo)
        VALUES(?,?,?,?,?,?,?,?)
        ON CONFLICT(item,lot,location) DO UPDATE SET
          qty = qty + excluded.qty,
          item_name = excluded.item_name,
          spec = excluded.spec,
          brand = excluded.brand,
          memo = excluded.memo,
          updated_at = datetime('now','localtime')
        """, (item,item_name,lot,spec,brand,location,qty,memo))
        db.execute("INSERT INTO history(type,item,item_name,lot,spec,brand,qty,src,dst,memo) VALUES('IN',?,?,?,?,?,?,?,?,?)",
                   (item,item_name,lot,spec,brand,qty,"",location,memo))
    db.commit(); db.close()
    return RedirectResponse("/page/inbound", status_code=303)

@app.post("/api/outbound/excel")
async def api_outbound_excel(file: UploadFile = File(...)):
    data = await file.read()
    rows = _xlsx_rows(data)
    db = get_db()
    for d in rows:
        location = str(d.get("location","")).strip()
        item = str(d.get("item","")).strip()
        if not location or not item:
            continue
        item_name = str(d.get("item_name","") or "")
        lot = str(d.get("lot","") or "")
        spec = str(d.get("spec","") or "")
        brand = str(d.get("brand","") or "")
        memo = str(d.get("memo","") or "")
        qty = int(float(d.get("qty",0) or 0))
        if qty <= 0:
            continue
        db.execute("UPDATE stock SET qty = qty - ?, updated_at=datetime('now','localtime') WHERE item=? AND lot=? AND location=?",
                   (qty,item,lot,location))
        db.execute("INSERT INTO history(type,item,item_name,lot,spec,brand,qty,src,dst,memo) VALUES('OUT',?,?,?,?,?,?,?,?,?)",
                   (item,item_name,lot,spec,brand,qty,location,"",memo))
    db.commit(); db.close()
    return RedirectResponse("/page/outbound", status_code=303)

@app.post("/api/move/excel")
async def api_move_excel(file: UploadFile = File(...)):
    data = await file.read()
    rows = _xlsx_rows(data)
    db = get_db()
    for d in rows:
        src = str(d.get("src","")).strip()
        dst = str(d.get("dst","")).strip()
        item = str(d.get("item","")).strip()
        if not src or not dst or not item:
            continue
        item_name = str(d.get("item_name","") or "")
        lot = str(d.get("lot","") or "")
        spec = str(d.get("spec","") or "")
        brand = str(d.get("brand","") or "")
        memo = str(d.get("memo","") or "")
        qty = int(float(d.get("qty",0) or 0))
        if qty <= 0:
            continue
        db.execute("UPDATE stock SET qty = qty - ?, updated_at=datetime('now','localtime') WHERE item=? AND lot=? AND location=?",
                   (qty,item,lot,src))
        db.execute("""
        INSERT INTO stock(item,item_name,lot,spec,brand,location,qty,memo)
        VALUES(?,?,?,?,?,?,?,?)
        ON CONFLICT(item,lot,location) DO UPDATE SET
          qty = qty + excluded.qty,
          item_name = excluded.item_name,
          spec = excluded.spec,
          brand = excluded.brand,
          memo = excluded.memo,
          updated_at = datetime('now','localtime')
        """, (item,item_name,lot,spec,brand,dst,qty,memo))
        db.execute("INSERT INTO history(type,item,item_name,lot,spec,brand,qty,src,dst,memo) VALUES('MOVE',?,?,?,?,?,?,?,?,?)",
                   (item,item_name,lot,spec,brand,qty,src,dst,memo))
    db.commit(); db.close()
    return RedirectResponse("/page/move", status_code=303)


@app.get("/api/inventory")
def api_inventory(location: Optional[str]=None, item: Optional[str]=None):
    db=get_db()
    q="SELECT item,item_name,lot,spec,brand,location,qty,updated_at,memo FROM stock WHERE 1=1"
    args=[]
    if location:
        q+=" AND location LIKE ?"; args.append(f"%{location}%")
    if item:
        q+=" AND item LIKE ?"; args.append(f"%{item}%")
    q+=" ORDER BY location,item"
    rows=[dict(r) for r in db.execute(q,args).fetchall()]
    db.close()
    return {"rows": rows}


@app.get("/api/inventory.xlsx")
def api_inventory_xlsx(location: Optional[str]=None, item: Optional[str]=None):
    db=get_db()
    q="SELECT location,item,lot,spec,brand,qty,updated_at,memo FROM stock WHERE 1=1"
    args=[]
    if location:
        q+=" AND location LIKE ?"; args.append(f"%{location}%")
    if item:
        q+=" AND item LIKE ?"; args.append(f"%{item}%")
    q+=" ORDER BY location,item"
    rows=[dict(r) for r in db.execute(q,args).fetchall()]
    db.close()
    headers=["로케이션","품번","LOT","규격","브랜드","수량","업데이트","비고"]
    data_rows=[[r["location"], r["item"], r["lot"], r["spec"], r["brand"], r["qty"], r["updated_at"], r["memo"]] for r in rows]
    fn="inventory.xlsx"
    return _xlsx_response(fn, headers, data_rows)

@app.get("/api/history")
def api_history(limit:int=200):
    db=get_db()
    rows=[dict(r) for r in db.execute("SELECT * FROM history ORDER BY id DESC LIMIT ?", (limit,)).fetchall()]
    db.close()
    return {"rows": rows}


@app.get("/api/history.xlsx")
def api_history_xlsx(limit:int=200):
    db=get_db()
    rows=[dict(r) for r in db.execute("SELECT * FROM history ORDER BY id DESC LIMIT ?", (limit,)).fetchall()]
    db.close()
    headers=["시간","구분","로케이션(출발)","로케이션(도착)","품번","LOT","규격","브랜드","수량","비고"]
    data_rows=[[r["ts"], r["type"], r["src"], r["dst"], r["item"], r["lot"], r["spec"], r.get("brand",""), r["qty"], r.get("memo","")] for r in rows]
    fn="history.xlsx"
    return _xlsx_response(fn, headers, data_rows)

@app.post("/api/calendar")
def api_calendar(action: str = Form("add"), id: int = Form(0), date: str = Form(""), memo: str = Form(""), author: str = Form("")):
    db=get_db()
    if action == "add":
        db.execute("INSERT INTO calendar_memo(date,memo,author) VALUES(?,?,?)", (date, memo, author))
    elif action == "update":
        db.execute("UPDATE calendar_memo SET memo=?, author=?, updated_at=datetime('now','localtime') WHERE id=?", (memo, author, id))
    elif action == "delete":
        db.execute("DELETE FROM calendar_memo WHERE id=?", (id,))
    db.commit(); db.close()
    return RedirectResponse("/page/calendar?date="+date, status_code=303)

@app.get("/page/calendar/month", response_class=HTMLResponse)
def page_calendar_month(
    request: Request,
    year: int = Query(default=datetime.date.today().year),
    month: int = Query(default=datetime.date.today().month),
):
    """월간 달력: 간단 메모용 (칸당 최대 4줄 표시)"""
    import calendar

    db = get_db()
    cal = calendar.Calendar(6)  # 일요일 시작
    today = datetime.now().date()

    days = []
    for d in cal.itermonthdates(year, month):
        rows = db.execute(
            "SELECT memo FROM calendar_memo WHERE date=? ORDER BY id",
            (d.isoformat(),)
        ).fetchall()
        memos = [r[0] for r in rows][:4]  # ✅ 4줄만
        days.append(
            {
                "date": d,
                "memos": memos,
                "in_month": (d.month == month),
                "is_today": (d == today),
                "is_weekend": (d.weekday() >= 5),
            }
        )
    db.close()

    # 7일 단위로 주(week) 분리
    weeks = [days[i : i + 7] for i in range(0, len(days), 7)]

    prev_month = 12 if month == 1 else month - 1
    prev_year = year - 1 if month == 1 else year
    next_month = 1 if month == 12 else month + 1
    next_year = year + 1 if month == 12 else year

    return templates.TemplateResponse(
        "calendar_month.html",
        {
            "request": request,
            "weeks": weeks,
            "year": year,
            "month": month,
            "prev_year": prev_year,
            "prev_month": prev_month,
            "next_year": next_year,
            "next_month": next_month,
        },
    )

